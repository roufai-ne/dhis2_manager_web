"""
Service de génération de fichiers Excel avec formatage
=======================================================
Crée des fichiers Excel formatés avec styles, protection, etc.
"""

import logging
from typing import Dict, Optional
from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)


class ExcelService:
    """
    Service de génération et formatage de fichiers Excel
    """
    
    # Styles prédéfinis
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    DATA_FONT = Font(name='Calibri', size=10)
    HIDDEN_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    BORDER_THIN = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    def __init__(self):
        """Initialise le service Excel"""
        pass
    
    def create_template_excel(
        self,
        df: pd.DataFrame,
        filepath: str,
        metadata: Dict,
        protect_technical: bool = False
    ) -> str:
        """
        Crée un fichier Excel formaté à partir d'un DataFrame
        
        Args:
            df: DataFrame contenant les données
            filepath: Chemin du fichier à créer
            metadata: Métadonnées du template (nom dataset, période, etc.)
            protect_technical: Si True, cache/protège les colonnes techniques
            
        Returns:
            Chemin du fichier créé
        """
        logger.info(f"Création fichier Excel: {filepath}")
        
        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Données"
        
        # Ajouter les métadonnées en haut
        self._add_metadata_header(ws, metadata)
        
        # Ajouter les données
        start_row = 6  # Commencer après le header de métadonnées
        
        # Colonnes visibles selon le schéma demandé
        visible_cols = [
            'section',
            'dataElementName', 'dataElement',
            'orgUnitName', 'orgUnitCode', 'orgUnit',
            'categoryOptionComboName', 'categoryOptionCombo',
            'attributeOptionComboName', 'attributeOptionCombo',
            'period', 'value'
        ]
        technical_cols = []
        
        # Écrire les en-têtes
        col_idx = 1
        col_mapping = {}
        
        # Colonnes visibles
        for col in visible_cols:
            if col in df.columns:
                cell = ws.cell(row=start_row, column=col_idx, value=col)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.BORDER_THIN
                col_mapping[col] = col_idx
                col_idx += 1
        
        # Pas de colonnes techniques cachées dans ce schéma
        
        # Écrire les données
        for row_idx, (_, row_data) in enumerate(df.iterrows(), start=start_row + 1):
            for col_name, col_pos in col_mapping.items():
                cell = ws.cell(row=row_idx, column=col_pos, value=row_data[col_name])
                cell.font = self.DATA_FONT
                cell.border = self.BORDER_THIN
                
                # Alignement selon le type de colonne
                if col_name == 'value':
                    cell.alignment = Alignment(horizontal='right')
                    # Ajouter une validation de données (nombres uniquement)
                    dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1=0)
                    dv.error = 'Entrez un nombre valide'
                    dv.errorTitle = 'Valeur invalide'
                    ws.add_data_validation(dv)
                    dv.add(cell)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Ajuster les largeurs de colonnes
        self._adjust_column_widths(ws, df, col_mapping, visible_cols)
        
        # Ne pas cacher de colonnes; toutes visibles
        
        # Figer les volets (header)
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
        
        # Ajouter une feuille d'instructions
        self._add_instructions_sheet(wb, metadata)
        
        # Sauvegarder
        wb.save(filepath)
        logger.info(f"Fichier Excel créé: {filepath}")
        
        return filepath
    
    def _add_metadata_header(self, ws, metadata: Dict):
        """Ajoute un en-tête avec les métadonnées"""
        # Titre
        cell = ws.cell(row=1, column=1, value=f"Template DHIS2 - {metadata.get('dataset_name', 'Dataset')}")
        cell.font = Font(name='Calibri', size=14, bold=True, color="2E5090")
        
        # Période
        cell = ws.cell(row=2, column=1, value=f"Période: {metadata.get('period', 'N/A')}")
        cell.font = Font(name='Calibri', size=10, italic=True)
        
        # Statistiques
        stats_text = f"Organisations: {metadata.get('org_units', 0)} | Lignes: {metadata.get('total_rows', 0)}"
        cell = ws.cell(row=3, column=1, value=stats_text)
        cell.font = Font(name='Calibri', size=9, color="666666")
        
        # Ligne vide
        ws.row_dimensions[4].height = 5
        ws.row_dimensions[5].height = 3
    
    def _add_instructions_sheet(self, wb: Workbook, metadata: Dict):
        """Ajoute une feuille d'instructions"""
        ws = wb.create_sheet("Instructions")
        
        instructions = [
            ("Template DHIS2 - Mode d'emploi", Font(size=14, bold=True, color="2E5090")),
            ("", None),
            ("1. Comment remplir ce template ?", Font(size=12, bold=True)),
            ("   • Renseignez uniquement la colonne VALEUR", Font(size=10)),
            ("   • Les autres colonnes sont en lecture seule", Font(size=10)),
            ("   • Ne modifiez pas les colonnes cachées (techniques)", Font(size=10)),
            ("", None),
            ("2. Validation des données", Font(size=12, bold=True)),
            ("   • La colonne VALEUR accepte uniquement des nombres", Font(size=10)),
            ("   • Les valeurs négatives ne sont pas autorisées", Font(size=10)),
            ("", None),
            ("3. Après la saisie", Font(size=12, bold=True)),
            ("   • Enregistrez le fichier", Font(size=10)),
            ("   • Importez-le dans le Calculateur DHIS2", Font(size=10)),
            ("   • Le système générera le payload JSON", Font(size=10)),
            ("", None),
            ("Informations du template:", Font(size=12, bold=True, color="4472C4")),
            (f"   Dataset: {metadata.get('dataset_name', 'N/A')}", Font(size=10)),
            (f"   Période: {metadata.get('period', 'N/A')}", Font(size=10)),
            (f"   Organisations: {metadata.get('org_units', 0)}", Font(size=10)),
            (f"   Lignes générées: {metadata.get('total_rows', 0)}", Font(size=10)),
        ]
        
        for row_idx, (text, font) in enumerate(instructions, start=1):
            cell = ws.cell(row=row_idx, column=1, value=text)
            if font:
                cell.font = font
        
        # Ajuster la largeur
        ws.column_dimensions['A'].width = 60
    
    def _adjust_column_widths(self, ws, df: pd.DataFrame, col_mapping: Dict, visible_cols: list):
        """Ajuste automatiquement les largeurs de colonnes"""
        for col_name, col_pos in col_mapping.items():
            # Largeurs prédéfinies
            widths = {                'section': 25,                'dataElementName': 40,
                'dataElement': 22,
                'orgUnitName': 35,
                'orgUnit': 22,
                'categoryOptionComboName': 35,
                'categoryOptionCombo': 22,
                'attributeOptionComboName': 30,
                'attributeOptionCombo': 22,
                'period': 16,
                'VALEUR': 12
            }
            width = widths.get(col_name, 20)

            ws.column_dimensions[self._get_column_letter(col_pos)].width = width
    
    def _get_column_letter(self, col_idx: int) -> str:
        """Convertit un index de colonne (1-based) en lettre"""
        from openpyxl.utils import get_column_letter
        return get_column_letter(col_idx)
    
    def validate_excel_file(self, filepath: str) -> tuple[bool, Optional[str]]:
        """
        Valide qu'un fichier Excel est bien formé
        
        Args:
            filepath: Chemin du fichier
            
        Returns:
            Tuple (valide, message d'erreur)
        """
        try:
            wb = load_workbook(filepath, read_only=True)
            
            # Vérifier qu'il y a au moins une feuille
            if not wb.sheetnames:
                return False, "Le fichier ne contient aucune feuille"
            
            # Vérifier la feuille "Données"
            if "Données" not in wb.sheetnames:
                return False, "La feuille 'Données' est manquante"
            
            ws = wb["Données"]
            
            # Vérifier les colonnes requises
            required_cols = ['INFO_STRUCTURE', 'INFO_INDICATEUR', 'VALEUR']
            header_row = None
            
            # Trouver la ligne d'en-tête
            for row in ws.iter_rows(min_row=1, max_row=10):
                values = [cell.value for cell in row if cell.value]
                if any(col in values for col in required_cols):
                    header_row = row
                    break
            
            if not header_row:
                return False, "En-têtes de colonnes introuvables"
            
            wb.close()
            return True, None
            
        except Exception as e:
            return False, f"Erreur lors de la validation: {str(e)}"
    
    def read_template_data(self, filepath: str) -> pd.DataFrame:
        """
        Lit les données d'un template Excel
        
        Args:
            filepath: Chemin du fichier
            
        Returns:
            DataFrame avec les données
        """
        try:
            # Lire le fichier Excel
            df = pd.read_excel(filepath, sheet_name="Données", skiprows=5)
            
            logger.info(f"Données lues: {len(df)} lignes")
            return df
            
        except Exception as e:
            logger.error(f"Erreur lecture Excel: {e}")
            raise
